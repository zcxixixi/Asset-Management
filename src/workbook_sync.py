#!/usr/bin/env python3
"""Rigorous workbook synchronization for assets.xlsx.

This module provides fail-fast, schema-validated, backup-safe, atomic updates
for the core sheets used by the dashboard:
- Holdings
- Daily
- Chart
- Exec
"""

from __future__ import annotations

import shutil
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


class WorkbookSyncError(RuntimeError):
    """Raised when workbook schema or consistency checks fail."""


@dataclass
class DailyRow:
    row_idx: int
    date_str: str
    cash_usd: float
    gold_usd: float
    stocks_usd: float
    total_usd: float
    nav: float
    note: str


def _header_map(ws, header_row: int = 1) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue
        key = str(value).strip().lower()
        if key:
            mapping[key] = col
    return mapping


def _require_columns(ws, sheet_name: str, required: List[str]) -> Dict[str, int]:
    mapping = _header_map(ws)
    missing = [col for col in required if col not in mapping]
    if missing:
        raise WorkbookSyncError(f"{sheet_name} missing required columns: {missing}")
    return mapping


def _last_non_empty_row(ws, col_idx: int) -> int:
    for row in range(ws.max_row, 1, -1):
        value = ws.cell(row=row, column=col_idx).value
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return row
    return 1


def _coerce_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 30000 <= float(value) <= 60000:
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(value))).date()
    text = str(value).strip()
    if not text:
        return None
    text = text.split(" ")[0]
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def normalize_date_str(value: Any) -> str:
    parsed = _coerce_date(value)
    if parsed is None:
        text = str(value).strip()
        return text.split(" ")[0] if text else ""
    return parsed.isoformat()


def _create_backup(path: Path) -> Path:
    backup_dir = path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{path.stem}_{stamp}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def _atomic_save(workbook, target_path: Path) -> None:
    with tempfile.NamedTemporaryFile(
        dir=target_path.parent,
        prefix=f"{target_path.stem}.",
        suffix=".tmp.xlsx",
        delete=False,
    ) as tmp:
        temp_path = Path(tmp.name)
    try:
        workbook.save(temp_path)
        temp_path.replace(target_path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)


def _write_holdings(ws, sync_dt: datetime, holding_updates: List[Dict[str, Any]]) -> None:
    cols = _require_columns(
        ws,
        "Holdings",
        ["timestamp", "symbol", "quantity", "price_usd", "market_value_usd"],
    )
    ts_col = cols["timestamp"]
    sym_col = cols["symbol"]
    qty_col = cols["quantity"]
    price_col = cols["price_usd"]
    mv_col = cols["market_value_usd"]

    by_symbol = {
        str(item.get("symbol", "")).strip().upper(): item
        for item in holding_updates
        if str(item.get("symbol", "")).strip()
    }

    if not by_symbol:
        raise WorkbookSyncError("No holding updates were provided")

    ts_fmt = ws.cell(row=2, column=ts_col).number_format or "yyyy-mm-dd hh:mm:ss"
    qty_fmt = ws.cell(row=2, column=qty_col).number_format or "General"
    price_fmt = ws.cell(row=2, column=price_col).number_format or "#,##0.00"
    mv_fmt = ws.cell(row=2, column=mv_col).number_format or "#,##0.00"

    updated_symbols = set()
    for row in range(2, ws.max_row + 1):
        symbol_raw = ws.cell(row=row, column=sym_col).value
        symbol = str(symbol_raw).strip().upper() if symbol_raw is not None else ""
        if not symbol:
            continue

        item = by_symbol.get(symbol)
        if item is None:
            continue

        ts_cell = ws.cell(row=row, column=ts_col)
        ts_cell.value = sync_dt
        ts_cell.number_format = ts_fmt

        qty_cell = ws.cell(row=row, column=qty_col)
        qty_cell.value = float(item["quantity"])
        qty_cell.number_format = qty_fmt

        p_cell = ws.cell(row=row, column=price_col)
        p_cell.value = float(item["price_usd"])
        p_cell.number_format = price_fmt

        mv_cell = ws.cell(row=row, column=mv_col)
        mv_cell.value = float(item["market_value_usd"])
        mv_cell.number_format = mv_fmt

        updated_symbols.add(symbol)

    missing = sorted(set(by_symbol.keys()) - updated_symbols)
    if missing:
        raise WorkbookSyncError(f"Holdings symbols not found in sheet: {missing}")


def _read_daily_rows(ws) -> tuple[List[DailyRow], Dict[str, int], Dict[str, str]]:
    cols = _require_columns(
        ws,
        "Daily",
        ["date", "cash_usd", "gold_usd", "stocks_usd", "total_usd", "nav", "note"],
    )

    date_col = cols["date"]
    cash_col = cols["cash_usd"]
    gold_col = cols["gold_usd"]
    stocks_col = cols["stocks_usd"]
    total_col = cols["total_usd"]
    nav_col = cols["nav"]
    note_col = cols["note"]

    formats = {
        "date": ws.cell(row=2, column=date_col).number_format or "General",
        "cash": ws.cell(row=2, column=cash_col).number_format or "#,##0.00",
        "gold": ws.cell(row=2, column=gold_col).number_format or "#,##0.00",
        "stocks": ws.cell(row=2, column=stocks_col).number_format or "#,##0.00",
        "total": ws.cell(row=2, column=total_col).number_format or "#,##0.00",
        "nav": ws.cell(row=2, column=nav_col).number_format or "0.00",
        "note": ws.cell(row=2, column=note_col).number_format or "General",
    }

    last_row = _last_non_empty_row(ws, date_col)
    rows: List[DailyRow] = []
    for row in range(2, last_row + 1):
        raw_date = ws.cell(row=row, column=date_col).value
        date_str = normalize_date_str(raw_date)
        if not date_str:
            continue
        rows.append(
            DailyRow(
                row_idx=row,
                date_str=date_str,
                cash_usd=float(ws.cell(row=row, column=cash_col).value or 0),
                gold_usd=float(ws.cell(row=row, column=gold_col).value or 0),
                stocks_usd=float(ws.cell(row=row, column=stocks_col).value or 0),
                total_usd=float(ws.cell(row=row, column=total_col).value or 0),
                nav=float(ws.cell(row=row, column=nav_col).value or 0),
                note=str(ws.cell(row=row, column=note_col).value or ""),
            )
        )

    if not rows:
        raise WorkbookSyncError("Daily has no usable rows")

    return rows, cols, formats


def _validate_daily_integrity(rows: List[DailyRow]) -> None:
    seen = set()
    prev = None
    for item in rows:
        if item.date_str in seen:
            raise WorkbookSyncError(f"Daily has duplicate date: {item.date_str}")
        seen.add(item.date_str)

        d = datetime.strptime(item.date_str, "%Y-%m-%d").date()
        if prev and d < prev:
            raise WorkbookSyncError("Daily dates are not monotonic")
        prev = d

        total_from_parts = round(item.cash_usd + item.gold_usd + item.stocks_usd, 2)
        if abs(total_from_parts - round(item.total_usd, 2)) > 0.05:
            raise WorkbookSyncError(
                f"Daily total mismatch on {item.date_str}: {total_from_parts} vs {item.total_usd}"
            )
        if item.nav <= 0:
            raise WorkbookSyncError(f"Daily nav must be positive on {item.date_str}")


def _write_daily(
    ws,
    sync_dt: datetime,
    assets_grouped: Dict[str, float],
    total_balance: float,
) -> List[DailyRow]:
    rows, cols, fmts = _read_daily_rows(ws)
    _validate_daily_integrity(rows)

    date_col = cols["date"]
    cash_col = cols["cash_usd"]
    gold_col = cols["gold_usd"]
    stocks_col = cols["stocks_usd"]
    total_col = cols["total_usd"]
    nav_col = cols["nav"]
    note_col = cols["note"]

    last = rows[-1]
    last_date = datetime.strptime(last.date_str, "%Y-%m-%d").date()

    nav_ratio = (last.nav / last.total_usd) if last.total_usd > 0 else 0
    if nav_ratio <= 0:
        raise WorkbookSyncError("Cannot derive nav ratio from last Daily row")

    current_cash = float(assets_grouped.get("Cash USD", assets_grouped.get("USD Cash", 0.0)))
    current_gold = float(assets_grouped.get("Gold USD", 0.0))
    current_stocks = float(assets_grouped.get("US Stocks", 0.0))
    current_total = float(total_balance)
    current_nav = round(current_total * nav_ratio, 4)

    target_date = sync_dt.date()

    def write_row(row_idx: int, d: date, cash: float, gold: float, stocks: float, total: float, nav: float, note: str) -> None:
        ws.cell(row=row_idx, column=date_col, value=d.isoformat())
        ws.cell(row=row_idx, column=date_col).number_format = fmts["date"]
        ws.cell(row=row_idx, column=cash_col, value=round(cash, 2))
        ws.cell(row=row_idx, column=cash_col).number_format = fmts["cash"]
        ws.cell(row=row_idx, column=gold_col, value=round(gold, 2))
        ws.cell(row=row_idx, column=gold_col).number_format = fmts["gold"]
        ws.cell(row=row_idx, column=stocks_col, value=round(stocks, 2))
        ws.cell(row=row_idx, column=stocks_col).number_format = fmts["stocks"]
        ws.cell(row=row_idx, column=total_col, value=round(total, 2))
        ws.cell(row=row_idx, column=total_col).number_format = fmts["total"]
        ws.cell(row=row_idx, column=nav_col, value=round(nav, 4))
        ws.cell(row=row_idx, column=nav_col).number_format = fmts["nav"]
        ws.cell(row=row_idx, column=note_col, value=note)
        ws.cell(row=row_idx, column=note_col).number_format = fmts["note"]

    if target_date < last_date:
        raise WorkbookSyncError(
            f"Sync date {target_date.isoformat()} is older than Daily last date {last_date.isoformat()}"
        )

    # append missing days
    next_row = rows[-1].row_idx + 1
    if target_date > last_date:
        carry_cash = last.cash_usd
        carry_gold = last.gold_usd
        carry_stocks = last.stocks_usd
        carry_total = last.total_usd
        d = last_date + timedelta(days=1)
        while d <= target_date:
            if d == target_date:
                write_row(
                    next_row,
                    d,
                    current_cash,
                    current_gold,
                    current_stocks,
                    current_total,
                    current_nav,
                    "broker-sync",
                )
            else:
                carry_nav = round(carry_total * nav_ratio, 4)
                write_row(
                    next_row,
                    d,
                    carry_cash,
                    carry_gold,
                    carry_stocks,
                    carry_total,
                    carry_nav,
                    "carry",
                )
            next_row += 1
            d += timedelta(days=1)
    else:
        # same day: overwrite last row with latest market values
        write_row(
            rows[-1].row_idx,
            target_date,
            current_cash,
            current_gold,
            current_stocks,
            current_total,
            current_nav,
            "broker-sync",
        )

    # canonicalize date strings and rebuild typed rows from sheet
    refreshed_rows, _, _ = _read_daily_rows(ws)
    for item in refreshed_rows:
        ws.cell(row=item.row_idx, column=date_col, value=item.date_str)
        ws.cell(row=item.row_idx, column=date_col).number_format = fmts["date"]

    _validate_daily_integrity(refreshed_rows)
    return refreshed_rows


def _write_chart(ws, daily_rows: List[DailyRow]) -> None:
    cols = _require_columns(ws, "Chart", ["date", "nav"])
    date_col = cols["date"]
    nav_col = cols["nav"]

    date_fmt = ws.cell(row=2, column=date_col).number_format or "General"
    nav_fmt = ws.cell(row=2, column=nav_col).number_format or "0.00"

    last_row = _last_non_empty_row(ws, date_col)
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=date_col, value=None)
        ws.cell(row=row, column=nav_col, value=None)

    row_idx = 2
    for item in daily_rows:
        ws.cell(row=row_idx, column=date_col, value=item.date_str)
        ws.cell(row=row_idx, column=date_col).number_format = date_fmt
        ws.cell(row=row_idx, column=nav_col, value=round(item.nav, 4))
        ws.cell(row=row_idx, column=nav_col).number_format = nav_fmt
        row_idx += 1


def _write_exec(ws, sync_dt: datetime) -> None:
    b2 = ws["B2"]
    b2_fmt = b2.number_format or "yyyy-mm-dd hh:mm:ss"
    b2.value = sync_dt
    b2.number_format = b2_fmt

    for row in range(1, min(ws.max_row, 40) + 1):
        for col in range(1, min(ws.max_column, 20) + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and value.strip() == "更新时间":
                target = ws.cell(row=row, column=col + 1)
                target.value = sync_dt
                target.number_format = target.number_format or b2_fmt
                return


def _validate_chart_matches_daily(ws_chart, daily_rows: List[DailyRow]) -> None:
    cols = _require_columns(ws_chart, "Chart", ["date", "nav"])
    date_col = cols["date"]
    nav_col = cols["nav"]

    last = _last_non_empty_row(ws_chart, date_col)
    chart_rows = []
    for r in range(2, last + 1):
        d = normalize_date_str(ws_chart.cell(row=r, column=date_col).value)
        if not d:
            continue
        nav = float(ws_chart.cell(row=r, column=nav_col).value or 0)
        chart_rows.append((d, round(nav, 4)))

    daily_pairs = [(row.date_str, round(row.nav, 4)) for row in daily_rows]
    if chart_rows != daily_pairs:
        raise WorkbookSyncError("Chart rows do not match Daily date/nav timeline")


def sync_workbook(
    workbook_path: str,
    sync_dt: datetime,
    holding_updates: List[Dict[str, Any]],
    assets_grouped: Dict[str, float],
    total_balance: float,
) -> Dict[str, Any]:
    path = Path(workbook_path)
    if not path.exists():
        raise WorkbookSyncError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(path)
    required_sheets = ["Holdings", "Daily", "Chart", "Exec"]
    missing_sheets = [name for name in required_sheets if name not in wb.sheetnames]
    if missing_sheets:
        raise WorkbookSyncError(f"Workbook missing required sheets: {missing_sheets}")

    _write_holdings(wb["Holdings"], sync_dt, holding_updates)
    daily_rows = _write_daily(wb["Daily"], sync_dt, assets_grouped, total_balance)
    _write_chart(wb["Chart"], daily_rows)
    _write_exec(wb["Exec"], sync_dt)
    _validate_chart_matches_daily(wb["Chart"], daily_rows)

    backup_path = _create_backup(path)
    _atomic_save(wb, path)

    return {
        "backup_path": str(backup_path),
        "last_daily_date": daily_rows[-1].date_str,
        "daily_count": len(daily_rows),
    }
