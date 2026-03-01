#!/usr/bin/env python3
"""
Real-Time Asset Synchronization Script
- Reads holdings from assets.xlsx
- Calculates live portfolio values via yfinance
- Builds chart/performance payload
- Generates personalized advisor briefing from latest news + optional LLM
"""
import json
import os
import sys
from datetime import datetime, timezone
from typing import Any, Dict, List, Tuple

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook

try:
    from openai import OpenAI
except Exception:  # pragma: no cover
    OpenAI = None

INPUT_PATH = "assets.xlsx"
OUTPUT_PATH = "src/data.json"
REQUIRED_HOLDINGS_COLUMNS = {"symbol", "name", "quantity"}
REQUIRED_DAILY_COLUMNS = {"date", "nav"}
MAX_NEWS_ITEMS = 8


def safe_float(value: Any) -> float:
    if pd.isna(value) or value is None:
        return 0.0
    try:
        if isinstance(value, str):
            value = value.replace(",", "").replace("$", "").strip()
            if not value:
                return 0.0
            return float(value)
        return float(value)
    except Exception:
        return 0.0


def to_yf_symbol(symbol: str) -> str:
    symbol = str(symbol).strip().upper()
    if symbol.endswith(".US"):
        symbol = symbol.replace(".US", "")
    if symbol in {"GOLD.CN", "GOLD"} or "GOLD" in symbol:
        return "GC=F"
    return symbol


def normalize_asset_label(name: str) -> str:
    key = str(name).strip().lower()
    if "cash" in key or key in {"usd", "cash usd"}:
        return "Cash USD"
    if "gold" in key:
        return "Gold USD"
    if "stock" in key or "equity" in key or key in {"us stocks", "us stock"}:
        return "US Stocks"
    return str(name).strip() or "Portfolio"


def format_asset_label(name: str, symbol: str) -> str:
    sym_upper = str(symbol).upper()
    if "GOLD" in sym_upper or "GC=F" in sym_upper:
        return "Gold USD"
    if ".US" in sym_upper or sym_upper in {"NVDA", "TSLA", "QQQ", "SGOV"}:
        return "US Stocks"
    if "CASH" in sym_upper or sym_upper in {"USD", "USDT"}:
        return "Cash USD"
    return normalize_asset_label(name)


def get_realtime_price(symbol: str) -> float:
    if symbol.upper() == "CASH":
        return 1.0
    yf_symbol = to_yf_symbol(symbol)
    try:
        ticker = yf.Ticker(yf_symbol)
        fast_info = ticker.fast_info
        last_price = fast_info.get("last_price") if hasattr(fast_info, "get") else fast_info["last_price"]
        if last_price and last_price > 0:
            return float(last_price)

        history = ticker.history(period="5d")
        if not history.empty:
            return float(history["Close"].dropna().iloc[-1])
        raise ValueError(f"No valid market price for {yf_symbol}")
    except Exception as exc:
        print(f"Warning: Failed to fetch {symbol} ({yf_symbol}): {exc}")
        return 0.0


def parse_published_at(raw_value: Any) -> str:
    if raw_value is None:
        return "Unknown"
    try:
        if isinstance(raw_value, (int, float)):
            dt = datetime.fromtimestamp(float(raw_value), tz=timezone.utc)
            return dt.strftime("%Y-%m-%d %H:%M:%S UTC")
        text = str(raw_value).strip()
        if not text:
            return "Unknown"
        dt = pd.to_datetime(text, utc=True, errors="coerce")
        if pd.isna(dt):
            return "Unknown"
        return dt.strftime("%Y-%m-%d %H:%M:%S UTC")
    except Exception:
        return "Unknown"


def collect_portfolio_news(symbols: List[str], max_items: int = MAX_NEWS_ITEMS) -> List[Dict[str, str]]:
    news_items: List[Dict[str, str]] = []
    seen: set[str] = set()

    for raw_symbol in symbols:
        symbol = str(raw_symbol).strip().upper()
        if not symbol or symbol in {"CASH", "USD", "USDT"}:
            continue

        yf_symbol = to_yf_symbol(symbol)
        try:
            ticker = yf.Ticker(yf_symbol)
            raw_news = ticker.news or []
        except Exception as exc:
            print(f"Warning: Failed to fetch news for {symbol}: {exc}")
            continue

        for item in raw_news:
            title = str(item.get("title", "")).strip()
            url = str(item.get("link") or item.get("url") or "").strip()
            if not title:
                continue

            unique_key = url or title
            if unique_key in seen:
                continue
            seen.add(unique_key)

            publisher = str(item.get("publisher") or item.get("source") or "Unknown").strip()
            published_at = parse_published_at(item.get("providerPublishTime") or item.get("published"))
            summary = str(item.get("summary") or "").strip()

            news_items.append(
                {
                    "symbol": symbol,
                    "title": title,
                    "publisher": publisher,
                    "published_at": published_at,
                    "url": url,
                    "summary": summary,
                }
            )

    def sort_key(entry: Dict[str, str]) -> pd.Timestamp:
        dt = pd.to_datetime(entry.get("published_at", ""), utc=True, errors="coerce")
        if pd.isna(dt):
            return pd.Timestamp(0, tz="UTC")
        return dt

    news_items.sort(key=sort_key, reverse=True)
    return news_items[:max_items]


def format_pct(value: float) -> str:
    return f"{value:+.2f}%"


def build_rule_based_briefing(
    assets: List[Dict[str, str]],
    holdings: List[Dict[str, str]],
    perf_7d: float,
    news_items: List[Dict[str, str]],
) -> Dict[str, Any]:
    assets_map = {item["label"]: safe_float(item["value"]) for item in assets}
    total_value = sum(assets_map.values()) or 1.0

    cash_pct = (assets_map.get("Cash USD", 0.0) / total_value) * 100
    gold_pct = (assets_map.get("Gold USD", 0.0) / total_value) * 100
    stocks_pct = (assets_map.get("US Stocks", 0.0) / total_value) * 100

    headline = "Portfolio Pulse: Balanced but Event-Sensitive"
    if news_items:
        first = news_items[0]
        headline = f"Portfolio Pulse: {first['symbol']} drives latest market narrative"

    macro_summary = (
        f"7-day NAV move is {format_pct(perf_7d)}. Current allocation is "
        f"Cash {cash_pct:.1f}%, Gold {gold_pct:.1f}%, US Stocks {stocks_pct:.1f}%. "
        "Portfolio remains diversified, but near-term volatility can be elevated around headline-heavy sessions."
    )

    suggestions: List[Dict[str, str]] = []

    if cash_pct >= 20:
        suggestions.append(
            {
                "asset": "Cash USD",
                "action": "Deploy gradually",
                "rationale": "Cash allocation is relatively high. Consider phased entries to reduce timing risk instead of one-time deployment.",
            }
        )
    else:
        suggestions.append(
            {
                "asset": "Cash USD",
                "action": "Maintain reserve",
                "rationale": "Cash buffer is moderate; keeping some dry powder supports flexibility during volatility spikes.",
            }
        )

    if perf_7d < -1:
        suggestions.append(
            {
                "asset": "US Stocks",
                "action": "Rebalance carefully",
                "rationale": "Recent NAV drawdown suggests tightening risk controls and avoiding oversized directional bets.",
            }
        )
    else:
        suggestions.append(
            {
                "asset": "US Stocks",
                "action": "Hold core exposure",
                "rationale": "Momentum is not deteriorating sharply. Keep core equity allocation while reviewing valuation-sensitive names.",
            }
        )

    if gold_pct > 10:
        suggestions.append(
            {
                "asset": "Gold USD",
                "action": "Keep hedge",
                "rationale": "Gold weight already provides macro hedge coverage against rate and risk-off uncertainty.",
            }
        )

    risks = [
        "Short-term market reactions to macro headlines can exceed fundamentals.",
        "Single-name concentration risk can amplify drawdowns.",
    ]

    verdict = "Keep diversified core positions and adjust incrementally, not aggressively."

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source": "rule-based",
        "headline": headline,
        "macro_summary": macro_summary,
        "verdict": verdict,
        "suggestions": suggestions[:4],
        "risks": risks,
        "news_context": news_items,
        "disclaimer": "Informational only, not financial advice.",
    }


def sanitize_briefing(raw: Any, fallback: Dict[str, Any], news_items: List[Dict[str, str]], source: str) -> Dict[str, Any]:
    if not isinstance(raw, dict):
        raw = {}

    headline = str(raw.get("headline") or fallback["headline"]).strip()
    macro_summary = str(raw.get("macro_summary") or fallback["macro_summary"]).strip()
    verdict = str(raw.get("verdict") or fallback["verdict"]).strip()

    risks: List[str] = []
    for risk in raw.get("risks", []):
        text = str(risk).strip()
        if text:
            risks.append(text)
    if not risks:
        risks = fallback["risks"]

    suggestions: List[Dict[str, str]] = []
    for item in raw.get("suggestions", []):
        if not isinstance(item, dict):
            continue
        asset = normalize_asset_label(item.get("asset") or "Portfolio")
        action = str(item.get("action") or "Hold").strip()
        rationale = str(item.get("rationale") or "No rationale provided.").strip()
        if not rationale:
            continue
        suggestions.append({"asset": asset, "action": action, "rationale": rationale})

    if not suggestions:
        suggestions = fallback["suggestions"]

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source": source,
        "headline": headline,
        "macro_summary": macro_summary,
        "verdict": verdict,
        "suggestions": suggestions[:4],
        "risks": risks[:5],
        "news_context": news_items,
        "disclaimer": "Informational only, not financial advice.",
    }


def generate_llm_briefing(
    assets: List[Dict[str, str]],
    holdings: List[Dict[str, str]],
    perf_7d: float,
    news_items: List[Dict[str, str]],
) -> Dict[str, Any]:
    fallback = build_rule_based_briefing(assets, holdings, perf_7d, news_items)

    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key or OpenAI is None:
        return fallback

    model = os.getenv("OPENAI_MODEL", "gpt-4.1-mini").strip() or "gpt-4.1-mini"

    payload = {
        "portfolio_assets": assets,
        "holdings": holdings,
        "perf_7d_pct": round(perf_7d, 4),
        "news_items": news_items,
    }

    system_prompt = (
        "You are a cautious portfolio analyst. "
        "Use only the provided portfolio and news context. "
        "Return strict JSON with keys: headline, macro_summary, verdict, suggestions, risks. "
        "suggestions must be an array of objects with keys: asset, action, rationale. "
        "Keep language concise, concrete, and risk-aware. Do not promise returns."
    )

    try:
        client = OpenAI(api_key=api_key)
        completion = client.chat.completions.create(
            model=model,
            temperature=0.2,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
            ],
        )
        content = completion.choices[0].message.content or "{}"
        parsed = json.loads(content)
        return sanitize_briefing(parsed, fallback, news_items, source=f"llm:{model}")
    except Exception as exc:
        print(f"Warning: LLM briefing failed, using fallback: {exc}")
        return fallback


def suggestion_to_insight_type(action: str) -> str:
    action_lower = str(action).lower()
    if any(word in action_lower for word in ["reduce", "trim", "rebalance", "tighten", "hedge", "defensive"]):
        return "warning"
    if any(word in action_lower for word in ["add", "deploy", "accumulate", "hold", "increase", "keep"]):
        return "opportunity"
    return "neutral"


def briefing_to_insights(briefing: Dict[str, Any]) -> List[Dict[str, str]]:
    suggestions = briefing.get("suggestions", []) if isinstance(briefing, dict) else []
    insights: List[Dict[str, str]] = []

    for item in suggestions[:3]:
        if not isinstance(item, dict):
            continue
        asset = normalize_asset_label(item.get("asset") or "Portfolio")
        action = str(item.get("action") or "").strip()
        rationale = str(item.get("rationale") or "").strip()
        if not rationale:
            continue
        insight_type = suggestion_to_insight_type(action)
        text = rationale
        insights.append({"type": insight_type, "asset": asset, "text": text})

    if not insights:
        insights = [
            {
                "type": "neutral",
                "asset": "Portfolio",
                "text": "No strong signal detected; maintain diversification and review positions on schedule.",
            }
        ]
    return insights


def write_sync_timestamp_to_excel(sync_dt: datetime) -> None:
    """
    Keep workbook timestamps aligned with JSON last_updated.
    Updates:
    - Holdings.timestamp for active holding rows
    - Exec!B2 (data time)
    - Exec cell right to label "更新时间" when present
    """
    if not os.path.exists(INPUT_PATH):
        return

    try:
        workbook = load_workbook(INPUT_PATH)
        updated = False
        default_fmt = "yyyy-mm-dd hh:mm:ss"

        if "Holdings" in workbook.sheetnames:
            ws = workbook["Holdings"]
            header_map = {
                str(ws.cell(row=1, column=col).value).strip().lower(): col
                for col in range(1, ws.max_column + 1)
            }
            ts_col = header_map.get("timestamp")
            symbol_col = header_map.get("symbol")
            if ts_col and symbol_col:
                base_fmt = ws.cell(row=2, column=ts_col).number_format or default_fmt
                for row in range(2, ws.max_row + 1):
                    symbol = ws.cell(row=row, column=symbol_col).value
                    if symbol is None or str(symbol).strip() == "":
                        continue
                    ts_cell = ws.cell(row=row, column=ts_col)
                    ts_cell.value = sync_dt
                    ts_cell.number_format = base_fmt
                    updated = True

        if "Exec" in workbook.sheetnames:
            ws = workbook["Exec"]
            b2 = ws["B2"]
            b2_fmt = b2.number_format or default_fmt
            b2.value = sync_dt
            b2.number_format = b2_fmt
            updated = True

            found_update_label = False
            for row in range(1, min(ws.max_row, 40) + 1):
                for col in range(1, min(ws.max_column, 20) + 1):
                    value = ws.cell(row=row, column=col).value
                    if isinstance(value, str) and value.strip() == "更新时间":
                        target = ws.cell(row=row, column=col + 1)
                        target.number_format = target.number_format or b2_fmt
                        target.value = sync_dt
                        found_update_label = True
                        updated = True
                        break
                if found_update_label:
                    break

        if updated:
            workbook.save(INPUT_PATH)
            print("Updated Excel timestamp fields in assets.xlsx")
    except Exception as exc:
        print(f"Warning: Failed to write sync timestamp back to Excel: {exc}")


def extract_data() -> Dict[str, Any]:
    if not os.path.exists(INPUT_PATH):
        raise FileNotFoundError(f"{INPUT_PATH} not found in repository root")

    print("Reading local Excel Holdings (Broker Export Format)...")
    df_holdings = pd.read_excel(INPUT_PATH, sheet_name="Holdings")
    df_holdings = df_holdings.dropna(how="all")
    df_holdings.columns = [str(c).strip().lower() for c in df_holdings.columns]
    missing_holdings_columns = REQUIRED_HOLDINGS_COLUMNS - set(df_holdings.columns)
    if missing_holdings_columns:
        raise ValueError(f"Holdings sheet missing columns: {sorted(missing_holdings_columns)}")

    assets_grouped: Dict[str, float] = {}
    total_balance = 0.0
    live_holdings: List[Dict[str, str]] = []
    tracked_symbols: List[str] = []

    print("\nFetching real-time market data via yfinance...")
    for _, row in df_holdings.iterrows():
        if pd.isna(row.get("symbol")):
            continue

        symbol = str(row.get("symbol", "CASH")).strip()
        name = str(row.get("name", "Unknown")).strip()
        qty = safe_float(row.get("quantity"))

        price = get_realtime_price(symbol)
        usd_value = qty * price
        total_balance += usd_value

        print(f"[{symbol}] {qty} units @ ${price:,.2f} = ${usd_value:,.2f}")

        category_label = format_asset_label(name, symbol)
        assets_grouped[category_label] = assets_grouped.get(category_label, 0.0) + usd_value

        if symbol.upper() not in ("CASH", "USD") and usd_value > 0:
            live_holdings.append(
                {
                    "symbol": symbol,
                    "name": name,
                    "qty": str(qty) if qty != int(qty) else str(int(qty)),
                    "value": f"{usd_value:,.2f}",
                }
            )
        tracked_symbols.append(symbol)

    final_assets = [
        {"label": "Cash USD", "value": f"{assets_grouped.get('Cash USD', assets_grouped.get('USD Cash', 0.0)):,.2f}"},
        {"label": "Gold USD", "value": f"{assets_grouped.get('Gold USD', 0.0):,.2f}"},
        {"label": "US Stocks", "value": f"{assets_grouped.get('US Stocks', 0.0):,.2f}"},
    ]

    df_daily = pd.read_excel(INPUT_PATH, sheet_name="Daily")
    df_daily = df_daily.dropna(how="all", axis=1).dropna(how="all", axis=0)
    df_daily.columns = df_daily.columns.astype(str).str.strip().str.lower()
    missing_daily_columns = REQUIRED_DAILY_COLUMNS - set(df_daily.columns)
    if missing_daily_columns:
        raise ValueError(f"Daily sheet missing columns: {sorted(missing_daily_columns)}")

    nav_series = df_daily["nav"].apply(safe_float).dropna()
    if nav_series.empty:
        raise ValueError("Daily.nav has no usable values")

    current_nav = nav_series.iloc[-1]
    nav_7d_ago = nav_series.iloc[-8] if len(nav_series) >= 8 else nav_series.iloc[0]
    perf_7d = ((current_nav / nav_7d_ago) - 1) * 100 if nav_7d_ago > 0 else 0.0

    chart_data = [
        {"date": str(row["date"]).split(" ")[0], "value": safe_float(row["nav"])}
        for _, row in df_daily.iterrows()
    ]
    if not chart_data:
        raise ValueError("chart_data generated empty from Daily sheet")

    print("\nCollecting latest portfolio-related news...")
    news_items = collect_portfolio_news(tracked_symbols, max_items=MAX_NEWS_ITEMS)
    print(f"Collected {len(news_items)} news items")

    print("Generating personalized advisor briefing...")
    advisor_briefing = generate_llm_briefing(final_assets, live_holdings, perf_7d, news_items)
    insights = briefing_to_insights(advisor_briefing)

    sync_now = datetime.now()
    last_updated_text = sync_now.strftime("%Y-%m-%d %H:%M:%S")
    write_sync_timestamp_to_excel(sync_now)

    response = {
        "assets": final_assets,
        "holdings": live_holdings,
        "chart_data": chart_data,
        "total_balance": f"{total_balance:,.2f}",
        "last_updated": last_updated_text,
        "insights": insights,
        "advisor_briefing": advisor_briefing,
        "performance": {"1d": "Live", "summary": "Broker Export Integration"},
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as file:
        json.dump(response, file, indent=2, ensure_ascii=False)

    print(f"\n✅ Pushed live calculated data to {OUTPUT_PATH}. Total NAV: ${total_balance:,.2f}")
    return response


if __name__ == "__main__":
    try:
        extract_data()
    except Exception as exc:
        print(f"Error orchestrating pipeline: {exc}", file=sys.stderr)
        raise
